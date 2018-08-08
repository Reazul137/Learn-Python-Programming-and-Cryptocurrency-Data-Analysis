import requests
from openpyxl import Workbook
import datetime

# This is the code for the Public Api from Coinmarketcap
import requests
from openpyxl import Workbook
import datetime

api = 'https://api.coinmarketcap.com/v2/ticker/'

raw_data = requests.get(api).json()
data = raw_data['data']
today = datetime.date.today()

file = Workbook()
sheet = file.create_sheet(str(today),0)

for currency in data:
    sheet.append(data[currency]['name'],
                 data[currency]['quote']['USD']['price'],
                 data[currency]['quote']['USD']['percent_change_1h'],
                 data[currency]['quote']['USD']['percent_change_24h'],
                 data[currency]['quote']['USD']['percent_change_7d']])

file.save("Data Alalysis.xlsx")

# This is the code for the Private Api from Coinmarketcap

# If you want to use this version remove the (3 quotes) ''' from the beginning and the end of the version
# and delete the Public Api version
# The 3 quotes mean that the code between them you not be executed

''' #Beginning of comment

key = 'YOUR COINMARKETCAP PRIVATE API KEY'

api = 'https://pro-api.coinmarketcap.com/v1/cryptocurrency/listings/latest?CMC_PRO_API_KEY='
api += key

raw_data = requests.get(api).json()
data = raw_data['data']
today = datetime.date.today()

file = Workbook()
sheet = file.create_sheet(str(today),0)

for currency in data:
    sheet.append([currency['name'],
               currency['quote']['USD']['price'],
               currency['quote']['USD']['percent_change_1h'],
               currency['quote']['USD']['percent_change_24h'],
               currency['quote']['USD']['percent_change_7d']])


file.save("Data Alalysis.xlsx")
 #End of Comment '''